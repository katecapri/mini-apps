"""
The script converts data from json format to xlsx tables.
If the file has several levels of nesting, the first level is distributed across pages.
The title of the json file in line 158. The title for the xlsx - in 190.
"""

import json

import openpyxl


def modify_first_row(ws):  # In the case of nesting, distributes the column names over several rows.
    max_col_on_sheet = ws.max_column
    names_of_columns = {}
    max_len_of_col = 0
    for j in range(1, max_col_on_sheet + 1):
        names_of_columns[j] = ws.cell(row=1, column=j).value.split('(')
        ws.cell(row=1, column=j).value = ''
        if len(names_of_columns[j]) > max_len_of_col:
            max_len_of_col = len(names_of_columns[j])
    if max_col_on_sheet > 1:
        ws.insert_rows(2, max_len_of_col - 1)
    for col in range(1, max_col_on_sheet + 1):
        for row in range(max_len_of_col, max_len_of_col - len(names_of_columns[col]), -1):
            ws.cell(row=row, column=col).value = names_of_columns[col][max_len_of_col - row].replace(')', '')
    flag_of_repeat_name = False
    dict_of_repeats = {}
    for row in range(1, max_len_of_col):
        for col in range(1, max_col_on_sheet):
            if not flag_of_repeat_name:
                if ws.cell(row=row, column=col).value == ws.cell(row=row, column=col + 1).value:
                    flag_of_repeat_name = True
                    dict_of_repeats[(row, col)] = 2
                    row_of_repeat = row
                    col_of_repeat = col
            else:
                if ws.cell(row=row, column=col).value == ws.cell(row=row, column=col + 1).value:
                    dict_of_repeats[(row_of_repeat, col_of_repeat)] += 1
                else:
                    flag_of_repeat_name = False
        flag_of_repeat_name = False
    for x, y in dict_of_repeats.keys():
        start = ws.cell(row=x, column=y).coordinate
        end = ws.cell(row=x, column=y + dict_of_repeats[x, y] - 1).coordinate
        ws.merge_cells(f'{start}:{end}')
    return max_len_of_col


def get_list_of_empty_cols():  # Getting a list of columns without data.
    count_cells_in_col = {}
    for row in sheet.keys():
        for col in sheet[row].keys():
            if type(sheet[row][col]) == str:
                if len(sheet[row][col]) > 0:
                    if col not in count_cells_in_col.keys():
                        count_cells_in_col[col] = 0
                    else:
                        count_cells_in_col[col] += 1
            else:
                if col not in count_cells_in_col.keys():
                    count_cells_in_col[col] = 0
                else:
                    count_cells_in_col[col] += 1
    list_of_empty_cols = []
    for col in count_cells_in_col.keys():
        if count_cells_in_col[col] == 0:
            list_of_empty_cols.append(col)
    return list_of_empty_cols


def save_to_excel_sheet():  # Transferring data from python dictionary to Excel sheet with the removal of empty lines
    list_of_empty_cols = get_list_of_empty_cols()
    sheet_book = book.active
    i = 1
    for line in sheet.values():
        for key, value in line.items():
            sheet_book.cell(row=i, column=key).value = str(value)
        i += 1
    for empty_col in list_of_empty_cols[::-1]:
        sheet_book.delete_cols(empty_col, 1)
    rows_in_hat = modify_first_row(sheet_book)
    cell_to_fix = sheet_book.cell(row=1 + rows_in_hat, column=1).coordinate
    sheet_book.freeze_panes = cell_to_fix


def deploy_list(duplicate_info, col_with_list):
    global max_row, max_col
    cell_for_deploy = duplicate_info[col_with_list]
    for cell in cell_for_deploy:
        sheet[max_row + 1] = {}
        for col2 in duplicate_info.keys():
            sheet[max_row + 1][col2] = duplicate_info[col2] if col2 != col_with_list else cell
        max_row += 1


def deploy_dict(row_with_dict, col_with_dict):
    global max_row, max_col
    cell_for_deploy = sheet[row_with_dict][col_with_dict]
    for key, value in cell_for_deploy.items():
        name_of_new_col = key + '(' + sheet[1][col_with_dict] + ')'
        if name_of_new_col in sheet[1].values():
            invert_headers = {value: key for key, value in sheet[1].items()}
            col_with_new_name = invert_headers[name_of_new_col]
            sheet[row_with_dict][col_with_new_name] = value
        else:
            sheet[1][max_col + 1] = name_of_new_col
            sheet[row_with_dict][max_col + 1] = value
            max_col += 1
    sheet[row_with_dict][col_with_dict] = ''


def get_number_of_lists_and_dicts():
    lists = 0
    dicts = 0
    for row in range(2, max_row + 1):
        if row in sheet.keys():
            for col in range(1, max_col + 1):
                if type(sheet.get(row).get(col)) == list:
                    lists += 1
                if type(sheet.get(row).get(col)) == dict:
                    dicts += 1
    return lists, dicts


def check_need_for_deploy():
    count_lists_on_sheet, count_dicts_on_sheet = get_number_of_lists_and_dicts()
    if count_lists_on_sheet == 0 and count_dicts_on_sheet == 0:
        return
    if count_dicts_on_sheet > 0:
        search_dicts()
    if count_lists_on_sheet > 0:
        for _ in range(1, count_lists_on_sheet + 1):
            search_list()
    check_need_for_deploy()


def search_list():
    rows = sheet.keys()
    for row in rows:
        for col in range(1, max_col + 1):
            if type(sheet.get(row).get(col)) == list:
                if len(sheet.get(row).get(col)) > 0:
                    duplicate_line = sheet.pop(row)
                    deploy_list(duplicate_line, col)
                else:
                    sheet[row][col] = ''
                return


def search_dicts():
    rows = sheet.keys()
    for row in rows:
        for col in range(1, max_col + 1):
            if type(sheet.get(row).get(col)) == dict:
                deploy_dict(row, col)


with open("{.json", "r", encoding='utf-8') as my_file:
    info_json = my_file.read()
    start_data = json.loads(info_json)
book = openpyxl.Workbook()
nesting_check = 0
for el in start_data.values():
    if type(el) == dict or type(el) == list:
        nesting_check += 1
if nesting_check > 1:
    sh = 0
    for el in start_data.keys():
        sheet = {1: {}, 2: {}}
        sheet[1][1] = el
        sheet[2][1] = start_data[el]
        max_col = 1
        max_row = 2
        book.create_sheet(index=sh, title=el)
        book.active = sh
        sh += 1
        check_need_for_deploy()
        save_to_excel_sheet()
else:
    sheet = {1: {}, 2: {}}
    max_col = 0
    for el in start_data.keys():
        max_col += 1
        sheet[1][max_col] = el
        sheet[2][max_col] = start_data[el]
    max_row = 2
    book.active = 0
    check_need_for_deploy()
    save_to_excel_sheet()
book.save('res.xlsx')
book.close()
